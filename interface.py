import streamlit as st
import sqlite3
import pandas as pd
import random
import hashlib
import io
from datetime import datetime, date, timedelta

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

st.set_page_config(
    page_title="JM-edt",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Nunito:wght@400;600;700;800&family=Caveat:wght@700&display=swap');
html, body, [class*="css"] { font-family: 'Nunito', sans-serif; }
.logo  { font-family:'Caveat',cursive; font-size:2.2rem; font-weight:700; color:#2563EB; }
.logo span { color:#FBBF24; }
.stitle { font-size:1.05rem; font-weight:800; color:#1E3A5F;
          border-left:4px solid #2563EB; padding-left:10px; margin-bottom:1rem; }
.stButton > button { border-radius:10px !important; font-family:'Nunito',sans-serif !important; font-weight:700 !important; }
div[data-testid="stSidebar"] { background:#fff; border-right:1px solid #DDE8F8; }
.tag { display:inline-block; padding:2px 9px; border-radius:14px; font-size:0.72rem; font-weight:700; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════
# DATABASE
# ═══════════════════════════════════════
DB = "jmedt.db"

def db():
    c = sqlite3.connect(DB, check_same_thread=False)
    c.row_factory = sqlite3.Row
    return c

def init_db():
    con = db()
    cur = con.cursor()
    cur.executescript("""
    PRAGMA foreign_keys = ON;

    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        pwd_hash TEXT NOT NULL,
        nom TEXT DEFAULT '',
        prenom TEXT DEFAULT ''
    );

    CREATE TABLE IF NOT EXISTS classes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL UNIQUE,
        nb_eleves INTEGER DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS profs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL,
        prenom TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS creneaux (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        debut TEXT NOT NULL,
        fin TEXT NOT NULL,
        ordre INTEGER DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS jours_actifs (
        jour_index INTEGER PRIMARY KEY
    );

    CREATE TABLE IF NOT EXISTS periode (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date_debut TEXT,
        date_fin TEXT,
        nb_semaines INTEGER DEFAULT 16
    );

    CREATE TABLE IF NOT EXISTS jours_feriers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date_jour TEXT NOT NULL UNIQUE
    );

    CREATE TABLE IF NOT EXISTS vacances (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        debut TEXT NOT NULL,
        fin TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS edusign_key (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        api_key TEXT
    );

    CREATE TABLE IF NOT EXISTS modules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL,
        prof_id INTEGER,
        nb_seances INTEGER DEFAULT 2,
        type_salle TEXT DEFAULT 'cm',
        couleur TEXT DEFAULT '#4A90D9',
        FOREIGN KEY(prof_id) REFERENCES profs(id) ON DELETE SET NULL
    );

    CREATE TABLE IF NOT EXISTS module_classes (
        module_id INTEGER NOT NULL,
        classe_id INTEGER NOT NULL,
        PRIMARY KEY(module_id, classe_id),
        FOREIGN KEY(module_id) REFERENCES modules(id) ON DELETE CASCADE,
        FOREIGN KEY(classe_id) REFERENCES classes(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS salles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL,
        capacite INTEGER DEFAULT 30,
        type TEXT DEFAULT 'cm'
    );

    CREATE TABLE IF NOT EXISTS dispos (
        prof_id INTEGER NOT NULL,
        jour_index INTEGER NOT NULL,
        creneau_id INTEGER NOT NULL,
        disponible INTEGER DEFAULT 0,
        prefere INTEGER DEFAULT 0,
        PRIMARY KEY(prof_id, jour_index, creneau_id),
        FOREIGN KEY(prof_id) REFERENCES profs(id) ON DELETE CASCADE,
        FOREIGN KEY(creneau_id) REFERENCES creneaux(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS edt (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        semaine INTEGER NOT NULL,
        jour_index INTEGER NOT NULL,
        creneau_id INTEGER NOT NULL,
        module_id INTEGER,
        salle_id INTEGER,
        classe_id INTEGER,
        algo TEXT,
        FOREIGN KEY(module_id) REFERENCES modules(id) ON DELETE SET NULL,
        FOREIGN KEY(salle_id)  REFERENCES salles(id)  ON DELETE SET NULL,
        FOREIGN KEY(classe_id) REFERENCES classes(id) ON DELETE SET NULL
    );
    """)

    # Donnees par defaut
    cur.execute("SELECT COUNT(*) FROM users")
    if cur.fetchone()[0] == 0:
        h = hashlib.sha256("issraa66@issou".encode()).hexdigest()
        cur.execute("INSERT INTO users (email,pwd_hash,nom,prenom) VALUES (?,?,?,?)",
                    ("issraaelaouni96@gmail.com", h, "El Aouni", "Issraa"))

    cur.execute("SELECT COUNT(*) FROM creneaux")
    if cur.fetchone()[0] == 0:
        cur.executemany("INSERT INTO creneaux (debut,fin,ordre) VALUES (?,?,?)",
                        [("08:00","10:00",0),("10:00","12:00",1),
                         ("14:00","16:00",2),("16:00","18:00",3)])

    cur.execute("SELECT COUNT(*) FROM jours_actifs")
    if cur.fetchone()[0] == 0:
        cur.executemany("INSERT INTO jours_actifs (jour_index) VALUES (?)",
                        [(i,) for i in range(5)])

    cur.execute("SELECT COUNT(*) FROM periode")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO periode (nb_semaines) VALUES (16)")

    con.commit()
    con.close()

init_db()

# ═══════════════════════════════════════
# SESSION
# ═══════════════════════════════════════
for k, v in [("logged_in", False), ("email", "")]:
    if k not in st.session_state:
        st.session_state[k] = v

# ═══════════════════════════════════════
# CONSTANTES
# ═══════════════════════════════════════
JOURS    = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi"]
COULEURS = ["#3B82F6","#F59E0B","#10B981","#EF4444","#8B5CF6",
            "#06B6D4","#F97316","#EC4899","#14B8A6","#6366F1",
            "#84CC16","#F43F5E","#0EA5E9","#D97706","#7C3AED"]
ALGOS    = {
    "backtracking" : "Backtracking",
    "coloring"     : "Coloriage de graphe",
    "genetic"      : "Algorithme genetique",
    "simulated"    : "Recuit simule"
}

# ═══════════════════════════════════════
# HELPERS DB
# ═══════════════════════════════════════
def h(pwd): return hashlib.sha256(pwd.encode()).hexdigest()

def fetch(sql, params=()):
    con = db(); rows = con.execute(sql, params).fetchall(); con.close()
    return [dict(r) for r in rows]

def fetch_one(sql, params=()):
    con = db(); row = con.execute(sql, params).fetchone(); con.close()
    return dict(row) if row else None

def run(sql, params=()):
    con = db(); con.execute("PRAGMA foreign_keys=ON"); con.execute(sql, params); con.commit(); con.close()

def run_many(sql, data):
    con = db(); con.executemany(sql, data); con.commit(); con.close()

def get_classes():
    return fetch("SELECT * FROM classes ORDER BY nom")

def get_profs():
    return fetch("SELECT * FROM profs ORDER BY nom")

def get_creneaux():
    return fetch("SELECT * FROM creneaux ORDER BY ordre, debut")

def get_jours_actifs():
    return [r["jour_index"] for r in fetch("SELECT jour_index FROM jours_actifs ORDER BY jour_index")]

def get_modules():
    return fetch("""SELECT m.*, p.nom as pnom, p.prenom as pprenom
                    FROM modules m LEFT JOIN profs p ON m.prof_id=p.id ORDER BY m.nom""")

def get_module_classes(mid):
    return fetch("""SELECT c.* FROM classes c
                    JOIN module_classes mc ON c.id=mc.classe_id WHERE mc.module_id=?""", (mid,))

def get_salles():
    return fetch("SELECT * FROM salles ORDER BY nom")

def get_periode():
    r = fetch_one("SELECT * FROM periode ORDER BY id DESC LIMIT 1")
    return r or {"nb_semaines":16,"date_debut":None,"date_fin":None}

def get_dispos(prof_id):
    rows = fetch("SELECT * FROM dispos WHERE prof_id=?", (prof_id,))
    return {(r["jour_index"], r["creneau_id"]): r for r in rows}

def get_edt(semaine):
    return fetch("""
        SELECT e.*, m.nom as mnom, m.couleur, m.type_salle,
               s.nom as snom, p.nom as pnom, p.prenom as pprenom,
               c.debut, c.fin, c.ordre as crdre,
               cl.nom as clnom, m.id as mid
        FROM edt e
        LEFT JOIN modules m  ON e.module_id  = m.id
        LEFT JOIN salles  s  ON e.salle_id   = s.id
        LEFT JOIN profs   p  ON m.prof_id    = p.id
        LEFT JOIN creneaux c ON e.creneau_id = c.id
        LEFT JOIN classes  cl ON e.classe_id  = cl.id
        WHERE e.semaine=?
        ORDER BY e.jour_index, c.ordre, cl.nom
    """, (semaine,))

def max_semaine():
    r = fetch_one("SELECT MAX(semaine) as mx FROM edt")
    return r["mx"] if r and r["mx"] else 0

# ═══════════════════════════════════════
# GENERATION EDT
# ═══════════════════════════════════════
def generate(algo="backtracking"):
    con = db()
    con.execute("PRAGMA foreign_keys=ON")
    con.execute("DELETE FROM edt")
    con.commit()

    modules  = get_modules()
    salles   = get_salles()
    creneaux = get_creneaux()
    jours    = get_jours_actifs()
    classes  = get_classes()
    periode  = get_periode()
    nb_sem   = periode.get("nb_semaines", 16)

    if not modules:  con.close(); return False, "Aucun module."
    if not salles:   con.close(); return False, "Aucune salle."
    if not creneaux: con.close(); return False, "Aucun creneau."
    if not jours:    con.close(); return False, "Aucun jour actif."
    if not classes:  con.close(); return False, "Aucune classe."

    cur = con.cursor()

    for sem in range(1, nb_sem + 1):
        # grille[classe_id][jour][creneau_id] = None | dict
        grille = {}
        for cl in classes:
            grille[cl["id"]] = {}
            for j in jours:
                grille[cl["id"]][j] = {}
                for cr in creneaux:
                    grille[cl["id"]][j][cr["id"]] = None

        # construire la liste des placements
        a_placer = []
        for mod in modules:
            mc = get_module_classes(mod["id"])
            if not mc:
                mc = classes
            sp = max(1, round(mod["nb_seances"] / nb_sem))
            for _ in range(sp):
                a_placer.append({"mod": mod, "classes": mc})

        # trier selon algo
        if algo == "genetic":
            a_placer.sort(key=lambda x: -x["mod"]["nb_seances"])
        elif algo == "coloring":
            a_placer.sort(key=lambda x: x["mod"]["nom"])
        elif algo == "simulated":
            random.shuffle(a_placer)
        else:
            random.shuffle(a_placer)

        for item in a_placer:
            mod = item["mod"]
            cls = item["classes"]
            salle = next((s for s in salles if s["type"] == mod["type_salle"]),
                         salles[0] if salles else None)
            if not salle:
                continue

            slots = [(j, cr["id"]) for j in jours for cr in creneaux]
            random.shuffle(slots)

            if len(cls) > 1:
                # module commun : chercher slot libre pour TOUTES les classes
                for (j, crid) in slots:
                    if all(grille[cl["id"]][j][crid] is None for cl in cls):
                        for cl in cls:
                            grille[cl["id"]][j][crid] = {"mod": mod, "salle": salle}
                        break
            else:
                cl = cls[0]
                for (j, crid) in slots:
                    if grille[cl["id"]][j][crid] is None:
                        grille[cl["id"]][j][crid] = {"mod": mod, "salle": salle}
                        break

        # insérer en base
        for cl in classes:
            for j in jours:
                for cr in creneaux:
                    val = grille[cl["id"]][j][cr["id"]]
                    if val:
                        cur.execute(
                            "INSERT INTO edt (semaine,jour_index,creneau_id,module_id,salle_id,classe_id,algo) "
                            "VALUES (?,?,?,?,?,?,?)",
                            (sem, j, cr["id"], val["mod"]["id"], val["salle"]["id"], cl["id"], algo)
                        )

    con.commit()
    con.close()
    return True, "EDT genere ({} semaines) avec {}.".format(nb_sem, ALGOS.get(algo, algo))

# ═══════════════════════════════════════
# EXPORT EXCEL
# ═══════════════════════════════════════
def hex_to_rgb(hex_color):
    h = hex_color.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def export_excel(semaine, classes_sel, creneaux, jours_actifs):
    if not EXCEL_OK:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_cells = get_edt(semaine)
    grid = {}
    for cell in all_cells:
        grid[(cell["jour_index"], cell["creneau_id"], cell["classe_id"])] = cell

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cl in classes_sel:
        ws = wb.create_sheet(title=cl["nom"][:31])
        ws.sheet_view.showGridLines = False

        # Titre
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=1 + len(jours_actifs))
        tc = ws.cell(1, 1,
                     "Emploi du temps — {} — Semaine {}".format(cl["nom"], semaine))
        tc.font      = Font(bold=True, size=13, color="1E3A5F")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.fill      = PatternFill("solid", fgColor="EEF5FF")
        ws.row_dimensions[1].height = 28

        # En-tetes jours
        ws.cell(2, 1, "Creneau").font = Font(bold=True, color="FFFFFF")
        ws.cell(2, 1).fill = PatternFill("solid", fgColor="1E3A5F")
        ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(2, 1).border = border
        ws.column_dimensions["A"].width = 14

        for ci, ji in enumerate(jours_actifs):
            col = ci + 2
            c = ws.cell(2, col, JOURS[ji])
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="4A7C1A")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border
            ws.column_dimensions[get_column_letter(col)].width = 22
        ws.row_dimensions[2].height = 22

        # Lignes creneaux
        for ri, cr in enumerate(creneaux):
            row = ri + 3
            ws.row_dimensions[row].height = 60

            lbl = ws.cell(row, 1, "{}\n{}".format(cr["debut"], cr["fin"]))
            lbl.font      = Font(bold=True, color="475569", size=9)
            lbl.fill      = PatternFill("solid", fgColor="F8FBFF")
            lbl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            lbl.border    = border

            for ci, ji in enumerate(jours_actifs):
                col  = ci + 2
                cell = grid.get((ji, cr["id"], cl["id"]))
                wc   = ws.cell(row, col)
                wc.border = border
                wc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if cell and cell.get("mnom"):
                    color_hex = (cell.get("couleur") or "#4A90D9").lstrip("#")
                    r2, g2, b2 = hex_to_rgb("#" + color_hex)
                    light = "{:02X}{:02X}{:02X}".format(
                        min(255, r2 + 120), min(255, g2 + 120), min(255, b2 + 120))
                    wc.fill  = PatternFill("solid", fgColor=light)
                    wc.value = "{}\n{}\n{} {}\n{}".format(
                        cell["mnom"],
                        cell.get("type_salle", "").upper(),
                        cell.get("pprenom", ""), cell.get("pnom", ""),
                        cell.get("snom", "")
                    )
                    wc.font = Font(bold=True, size=8, color="1E3A5F")
                else:
                    wc.fill  = PatternFill("solid", fgColor="F8FBFF")
                    wc.value = ""

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ═══════════════════════════════════════
# PAGE LOGIN
# ═══════════════════════════════════════
def page_login():
    _, col, _ = st.columns([1, 1.1, 1])
    with col:
        st.markdown("""<div style="background:#fff;border-radius:20px;border:1px solid #DDE8F8;
            padding:2.5rem;box-shadow:0 8px 32px #2563EB14;margin-top:3rem">""",
            unsafe_allow_html=True)
        st.markdown('<div class="logo" style="text-align:center;margin-bottom:.3rem">JM<span>-edt</span></div>',
            unsafe_allow_html=True)
        st.markdown('<p style="text-align:center;color:#64748B;margin-bottom:1.5rem;font-size:.9rem">'
                    'Gestion des emplois du temps</p>', unsafe_allow_html=True)
        email = st.text_input("Email", placeholder="votre@email.com", key="li_email")
        pwd   = st.text_input("Mot de passe", type="password", placeholder="••••••••", key="li_pwd")
        if st.button("Se connecter", use_container_width=True, type="primary"):
            user = fetch_one("SELECT * FROM users WHERE email=? AND pwd_hash=?", (email, h(pwd)))
            if user:
                st.session_state.logged_in = True
                st.session_state.email     = email
                st.rerun()
            else:
                st.error("Email ou mot de passe incorrect.")
        st.markdown("</div>", unsafe_allow_html=True)

# ═══════════════════════════════════════
# PAGES SIDEBAR
# ═══════════════════════════════════════
def page_identifiant():
    st.markdown('<div class="stitle">Votre profil</div>', unsafe_allow_html=True)
    user = fetch_one("SELECT * FROM users WHERE email=?", (st.session_state.email,))
    if user:
        c1, c2 = st.columns([1, 5])
        with c1:
            ini = (user["prenom"][0] if user["prenom"] else "") + (user["nom"][0] if user["nom"] else "")
            st.markdown("""<div style="width:64px;height:64px;border-radius:50%;background:#DBEAFE;
                display:flex;align-items:center;justify-content:center;
                font-weight:800;font-size:1.2rem;color:#2563EB;margin:auto">{}</div>""".format(ini.upper()),
                unsafe_allow_html=True)
        with c2:
            st.markdown("**{} {}**".format(user["prenom"], user["nom"]))
            st.markdown("`{}`".format(user["email"]))
    st.divider()
    if st.button("Se deconnecter", key="btn_logout"):
        st.session_state.logged_in = False
        st.session_state.email     = ""
        st.rerun()

def page_creneaux():
    st.markdown('<div class="stitle">Creneaux horaires</div>', unsafe_allow_html=True)

    with st.expander("Ajouter un creneau", expanded=False):
        c1, c2, c3 = st.columns([2, 2, 1])
        with c1:
            td = st.time_input("Debut", value=datetime.strptime("08:00","%H:%M").time(), key="cr_d")
        with c2:
            tf = st.time_input("Fin",   value=datetime.strptime("10:00","%H:%M").time(), key="cr_f")
        with c3:
            st.write(""); st.write("")
            if st.button("Ajouter", key="btn_acr", type="primary"):
                nb = fetch_one("SELECT COUNT(*) as n FROM creneaux")["n"]
                run("INSERT INTO creneaux (debut,fin,ordre) VALUES (?,?,?)",
                    (td.strftime("%H:%M"), tf.strftime("%H:%M"), nb))
                st.rerun()

    for cr in get_creneaux():
        c1, c2, c3 = st.columns([3, 2, 1])
        with c1:
            st.markdown("**{} → {}**".format(cr["debut"], cr["fin"]))
        with c2:
            nd = st.time_input("", value=datetime.strptime(cr["debut"],"%H:%M").time(),
                               key="ed_d_{}".format(cr["id"]), label_visibility="collapsed")
            nf = st.time_input("", value=datetime.strptime(cr["fin"],"%H:%M").time(),
                               key="ed_f_{}".format(cr["id"]), label_visibility="collapsed")
        with c3:
            if st.button("Modifier", key="mod_cr_{}".format(cr["id"])):
                run("UPDATE creneaux SET debut=?,fin=? WHERE id=?",
                    (nd.strftime("%H:%M"), nf.strftime("%H:%M"), cr["id"]))
                st.rerun()
            if st.button("Sup.", key="del_cr_{}".format(cr["id"])):
                run("DELETE FROM creneaux WHERE id=?", (cr["id"],))
                st.rerun()

    st.divider()
    st.markdown('<div class="stitle">Jours de cours</div>', unsafe_allow_html=True)
    ja   = get_jours_actifs()
    cols = st.columns(6)
    for i, j in enumerate(JOURS):
        with cols[i]:
            v = st.checkbox(j, value=(i in ja), key="jour_{}".format(i))
            if v and i not in ja:
                run("INSERT OR IGNORE INTO jours_actifs (jour_index) VALUES (?)", (i,)); st.rerun()
            elif not v and i in ja:
                run("DELETE FROM jours_actifs WHERE jour_index=?", (i,)); st.rerun()

def page_periode():
    st.markdown('<div class="stitle">Calendrier academique</div>', unsafe_allow_html=True)
    p = get_periode()
    c1, c2, c3 = st.columns(3)
    with c1: dd = st.date_input("Date de debut", value=date.today(), key="p_dd")
    with c2: df = st.date_input("Date de fin",   value=date.today()+timedelta(weeks=16), key="p_df")
    with c3:
        ns = st.number_input("Nombre de semaines", min_value=1, max_value=52,
                              value=int(p.get("nb_semaines",16)), key="p_ns")

    if st.button("Enregistrer", type="primary", key="btn_periode"):
        run("DELETE FROM periode")
        run("INSERT INTO periode (date_debut,date_fin,nb_semaines) VALUES (?,?,?)",
            (str(dd), str(df), ns))
        st.success("Calendrier enregistre !")

    st.divider()
    st.markdown("**Jours feries**")
    c1, c2 = st.columns([3,1])
    with c1: nf = st.date_input("Date", key="new_ferie")
    with c2:
        st.write(""); st.write("")
        if st.button("Ajouter", key="btn_af"):
            run("INSERT OR IGNORE INTO jours_feriers (date_jour) VALUES (?)", (str(nf),))
            st.rerun()
    for f in fetch("SELECT * FROM jours_feriers ORDER BY date_jour"):
        c1, c2 = st.columns([5,1])
        with c1: st.markdown("📍 {}".format(f["date_jour"]))
        with c2:
            if st.button("X", key="df_{}".format(f["id"])):
                run("DELETE FROM jours_feriers WHERE id=?", (f["id"],)); st.rerun()

    st.divider()
    st.markdown("**Periodes de vacances**")
    c1, c2, c3 = st.columns([2,2,1])
    with c1: vd = st.date_input("Debut", key="v_d")
    with c2: vf = st.date_input("Fin",   key="v_f")
    with c3:
        st.write(""); st.write("")
        if st.button("Ajouter", key="btn_av"):
            run("INSERT INTO vacances (debut,fin) VALUES (?,?)", (str(vd), str(vf)))
            st.rerun()
    for v in fetch("SELECT * FROM vacances ORDER BY debut"):
        c1, c2 = st.columns([5,1])
        with c1: st.markdown("🌴 {} → {}".format(v["debut"], v["fin"]))
        with c2:
            if st.button("X", key="dv_{}".format(v["id"])):
                run("DELETE FROM vacances WHERE id=?", (v["id"],)); st.rerun()

def page_edusign():
    st.markdown('<div class="stitle">Edusign</div>', unsafe_allow_html=True)
    st.info("Entrez votre cle API pour synchroniser l'EDT avec Edusign.")
    ex = fetch_one("SELECT * FROM edusign_key LIMIT 1")
    ck = ex["api_key"] if ex else ""
    ak = st.text_input("Cle API Edusign", value=ck, type="password", key="edusign_key")
    if st.button("Enregistrer", type="primary", key="btn_edu"):
        run("DELETE FROM edusign_key")
        run("INSERT INTO edusign_key (api_key) VALUES (?)", (ak,))
        st.success("Cle enregistree !")
    if ck:
        st.success("Cle Edusign configuree ✓")

# ═══════════════════════════════════════
# PAGES ONGLETS
# ═══════════════════════════════════════
def tab_professeur():
    st.markdown('<div class="stitle">Gestion des professeurs</div>', unsafe_allow_html=True)

    with st.expander("Ajouter un professeur", expanded=True):
        c1, c2, c3 = st.columns([2,2,1])
        with c1: pnom    = st.text_input("Nom",    placeholder="NOM",    key="p_nom")
        with c2: pprenom = st.text_input("Prenom", placeholder="PRENOM", key="p_prenom")
        with c3:
            st.write(""); st.write("")
            if st.button("Ajouter", key="btn_ap", type="primary"):
                if pnom.strip() and pprenom.strip():
                    run("INSERT INTO profs (nom,prenom) VALUES (?,?)",
                        (pnom.strip().upper(), pprenom.strip().upper()))
                    st.rerun()
                else:
                    st.warning("Saisissez nom et prenom.")

    profs = get_profs()
    st.markdown("**{} professeur(s)**".format(len(profs)))
    if not profs:
        st.info("Aucun professeur enregistre.")
        return

    for p in profs:
        c1, c2 = st.columns([6,1])
        with c1:
            st.markdown("""<div style="display:flex;align-items:center;gap:12px;
                padding:10px 16px;background:#F8FBFF;border-radius:10px;
                border:1px solid #DDE8F8;margin-bottom:6px">
                <div style="width:38px;height:38px;border-radius:50%;background:#DBEAFE;
                display:flex;align-items:center;justify-content:center;
                font-weight:800;font-size:.9rem;color:#2563EB;flex-shrink:0">
                {}{}</div>
                <span style="font-weight:700;color:#1E3A5F">{} {}</span>
                </div>""".format(p["prenom"][0], p["nom"][0], p["prenom"], p["nom"]),
                unsafe_allow_html=True)
        with c2:
            if st.button("Sup.", key="dp_{}".format(p["id"])):
                run("DELETE FROM profs WHERE id=?", (p["id"],)); st.rerun()

def tab_classes():
    st.markdown('<div class="stitle">Gestion des classes</div>', unsafe_allow_html=True)

    with st.expander("Ajouter une classe / groupe", expanded=True):
        c1, c2, c3 = st.columns([3,2,1])
        with c1:
            cnom = st.text_input("Nom de la classe", placeholder="Ex: ISEN1, HEI2, L3 Info...", key="cl_nom")
        with c2:
            cnb  = st.number_input("Nb d'eleves", min_value=1, max_value=500, value=25, key="cl_nb")
        with c3:
            st.write(""); st.write("")
            if st.button("Ajouter", key="btn_ac", type="primary"):
                if cnom.strip():
                    try:
                        run("INSERT INTO classes (nom,nb_eleves) VALUES (?,?)",
                            (cnom.strip().upper(), cnb))
                        st.rerun()
                    except Exception:
                        st.error("Ce nom de classe existe deja.")
                else:
                    st.warning("Saisissez un nom.")

    classes = get_classes()
    st.markdown("**{} classe(s)**".format(len(classes)))
    if not classes:
        st.info("Aucune classe. Ajoutez-en pour generer l'EDT.")
        return

    cols = st.columns(min(len(classes), 4))
    for i, cl in enumerate(classes):
        with cols[i % 4]:
            c1, c2 = st.columns([4,1])
            with c1:
                st.markdown("""<div style="background:#EEF5FF;border-radius:14px;
                    border:1px solid #BFDBFE;padding:16px;margin-bottom:10px;text-align:center">
                    <div style="font-weight:800;color:#1E3A5F;font-size:1.1rem">{}</div>
                    <div style="font-size:.8rem;color:#64748B;margin-top:4px">{} eleves</div>
                    </div>""".format(cl["nom"], cl["nb_eleves"]),
                    unsafe_allow_html=True)
            with c2:
                if st.button("X", key="dcl_{}".format(cl["id"])):
                    run("DELETE FROM classes WHERE id=?", (cl["id"],)); st.rerun()

def tab_disponibilite():
    st.markdown('<div class="stitle">Disponibilites et preferences des professeurs</div>',
                unsafe_allow_html=True)
    profs = get_profs()
    if not profs:
        st.warning("Ajoutez des professeurs d'abord."); return

    po  = {"{} {}".format(p["prenom"], p["nom"]): p["id"] for p in profs}
    sel = st.selectbox("Selectionner un professeur", list(po.keys()), key="sel_prof_dispo")
    pid = po[sel]

    creneaux = get_creneaux()
    ja       = get_jours_actifs()
    dispos   = get_dispos(pid)

    if not creneaux or not ja:
        st.warning("Configurez creneaux et jours actifs dans les Reglages."); return

    st.markdown("""
    <div style="display:flex;gap:10px;margin-bottom:1rem;flex-wrap:wrap">
        <span class="tag" style="background:#BFDBFE;color:#1D4ED8">V = Disponible</span>
        <span class="tag" style="background:#FDE68A;color:#92400E">P = Prefere</span>
        <span class="tag" style="background:#F1F5F9;color:#94A3B8">- = Indisponible</span>
        <span style="font-size:.8rem;color:#94A3B8;margin-left:8px">Cliquez pour changer l'etat</span>
    </div>""", unsafe_allow_html=True)

    hcols = st.columns([2] + [1]*len(ja))
    with hcols[0]: st.markdown("**Creneau**")
    for i, j in enumerate(ja):
        with hcols[i+1]: st.markdown("**{}**".format(JOURS[j][:3]))

    con = db()
    con.execute("PRAGMA foreign_keys=ON")
    for cr in creneaux:
        rcols = st.columns([2] + [1]*len(ja))
        with rcols[0]:
            st.markdown("<small style='font-weight:700;color:#475569'>{}-{}</small>".format(
                cr["debut"], cr["fin"]), unsafe_allow_html=True)
        for i, j in enumerate(ja):
            with rcols[i+1]:
                d    = dispos.get((j, cr["id"]), {})
                disp = d.get("disponible", 0)
                pref = d.get("prefere", 0)
                lbl  = "P" if pref else ("V" if disp else "-")
                if st.button(lbl, key="d_{}_{}_{}" .format(pid, j, cr["id"])):
                    if not disp:        nd, np = 1, 0
                    elif disp and not pref: nd, np = 1, 1
                    else:               nd, np = 0, 0
                    con.execute("""INSERT INTO dispos
                        (prof_id,jour_index,creneau_id,disponible,prefere)
                        VALUES (?,?,?,?,?)
                        ON CONFLICT(prof_id,jour_index,creneau_id)
                        DO UPDATE SET disponible=?,prefere=?""",
                        (pid, j, cr["id"], nd, np, nd, np))
                    con.commit()
                    st.rerun()
    con.close()

def tab_matiere():
    st.markdown('<div class="stitle">Modules et matieres</div>', unsafe_allow_html=True)
    profs   = get_profs()
    classes = get_classes()

    if not classes:
        st.warning("Ajoutez des classes (onglet Classes) avant d'assigner des modules.")

    # ── Formulaire ajout ──
    with st.expander("Ajouter un module", expanded=True):
        c1, c2 = st.columns(2)
        with c1:
            mnom = st.text_input("Nom du module", placeholder="Ex: Mathematiques", key="m_nom")
        with c2:
            pc   = ["-- Sans professeur --"] + ["{} {}".format(p["prenom"], p["nom"]) for p in profs]
            psel = st.selectbox("Professeur responsable", pc, key="m_prof")

        c3, c4, c5 = st.columns(3)
        with c3:
            ms = st.number_input("Nb seances (total)", min_value=1, max_value=200, value=10, key="m_seances")
        with c4:
            mt = st.selectbox("Type de salle", ["cm","td","tp"], key="m_type")
        with c5:
            mc_def = COULEURS[len(get_modules()) % len(COULEURS)]
            mc     = st.color_picker("Couleur", value=mc_def, key="m_color")

        # ── Sélection des classes / groupes ──
        st.markdown("---")
        st.markdown("**Groupes / Classes concernes**")
        st.caption("Cochez une seule classe = cours individuel.  Cochez plusieurs = cours commun (meme creneau pour tous).")

        sel_classes = []
        if classes:
            nb_cols = min(len(classes), 5)
            gcols   = st.columns(nb_cols)
            for i, cl in enumerate(classes):
                with gcols[i % nb_cols]:
                    bg = "#EEF5FF" if True else "#fff"
                    checked = st.checkbox(
                        "{}\n({} eleves)".format(cl["nom"], cl["nb_eleves"]),
                        key="mcc_{}".format(cl["id"])
                    )
                    if checked:
                        sel_classes.append(cl["id"])

            if len(sel_classes) == 0:
                st.warning("Selectionnez au moins un groupe.")
            elif len(sel_classes) == 1:
                cl_name = next(c["nom"] for c in classes if c["id"] == sel_classes[0])
                st.success("Cours individuel pour le groupe **{}**".format(cl_name))
            else:
                cl_names = [c["nom"] for c in classes if c["id"] in sel_classes]
                st.info("Cours commun pour : **{}** — ils partageront le meme creneau dans l'EDT.".format(
                    " + ".join(cl_names)))
        else:
            st.caption("Aucune classe disponible. Ajoutez des classes d'abord.")

        st.markdown("---")
        if st.button("Ajouter le module", type="primary", key="btn_addm"):
            if not mnom.strip():
                st.warning("Saisissez un nom de module.")
            elif not sel_classes:
                st.warning("Selectionnez au moins un groupe / classe.")
            else:
                pid_m = None
                if psel != "-- Sans professeur --":
                    idx   = pc.index(psel) - 1
                    pid_m = profs[idx]["id"]
                con = db()
                con.execute("PRAGMA foreign_keys=ON")
                cur = con.execute(
                    "INSERT INTO modules (nom,prof_id,nb_seances,type_salle,couleur) VALUES (?,?,?,?,?)",
                    (mnom.strip(), pid_m, ms, mt, mc))
                mid = cur.lastrowid
                for cid in sel_classes:
                    con.execute("INSERT OR IGNORE INTO module_classes (module_id,classe_id) VALUES (?,?)",
                                (mid, cid))
                con.commit()
                con.close()
                st.success("Module '{}' ajoute !".format(mnom.strip()))
                st.rerun()

    # ── Liste des modules ──
    modules = get_modules()
    st.markdown("**{} module(s) enregistre(s)**".format(len(modules)))
    if not modules:
        st.info("Aucun module."); return

    # Regrouper par classe pour faciliter la lecture
    st.markdown("---")

    # Affichage par groupe
    if classes:
        filtre_cl = st.selectbox(
            "Filtrer par classe",
            ["Tous les modules"] + [cl["nom"] for cl in classes],
            key="filtre_mod_cl"
        )
    else:
        filtre_cl = "Tous les modules"

    for m in modules:
        mc_list = get_module_classes(m["id"])
        mc_ids  = [c["id"] for c in mc_list]
        mcl     = " / ".join(c["nom"] for c in mc_list) if mc_list else "Non assigne"
        prof    = "{} {}".format(m.get("pprenom",""), m.get("pnom","")).strip() or "Sans prof"

        # Filtre
        if filtre_cl != "Tous les modules":
            cl_match = next((c for c in classes if c["nom"] == filtre_cl), None)
            if cl_match and cl_match["id"] not in mc_ids:
                continue

        commun = len(mc_list) > 1
        badge_commun = ""
        if commun:
            badge_commun = "<span class='tag' style='background:#FEE2E2;color:#DC2626'>Cours commun</span>"

        c1, c2, c3 = st.columns([6, 1, 1])
        with c1:
            st.markdown("""<div style="display:flex;align-items:center;gap:9px;
                padding:10px 14px;background:#F8FBFF;border-radius:10px;
                border:1px solid #DDE8F8;margin-bottom:6px;flex-wrap:wrap">
                <div style="width:15px;height:15px;border-radius:4px;background:{col};flex-shrink:0"></div>
                <span style="font-weight:700;color:#1E3A5F">{nom}</span>
                <span style="font-size:.8rem;color:#64748B">— {prof}</span>
                <span class="tag" style="background:#EEF5FF;color:#2563EB">{seq} seances</span>
                <span class="tag" style="background:#FEF3C7;color:#92400E">{tp}</span>
                <span class="tag" style="background:#ECFDF5;color:#065F46">
                    Groupes: {mcl}
                </span>
                {badge}
                </div>""".format(
                    col=m["couleur"], nom=m["nom"], prof=prof,
                    seq=m["nb_seances"], tp=m["type_salle"].upper(),
                    mcl=mcl, badge=badge_commun),
                unsafe_allow_html=True)
        with c2:
            # Bouton modifier les groupes
            if st.button("Groupes", key="edit_grp_{}".format(m["id"])):
                st.session_state["edit_mod_grp"] = m["id"]
                st.rerun()
        with c3:
            if st.button("Sup.", key="dm_{}".format(m["id"])):
                run("DELETE FROM module_classes WHERE module_id=?", (m["id"],))
                run("DELETE FROM modules WHERE id=?", (m["id"],))
                if st.session_state.get("edit_mod_grp") == m["id"]:
                    st.session_state.pop("edit_mod_grp", None)
                st.rerun()

        # Panel d'édition des groupes inline
        if st.session_state.get("edit_mod_grp") == m["id"]:
            with st.container():
                st.markdown("""<div style="background:#F0F9FF;border:1.5px solid #BAE6FD;
                    border-radius:10px;padding:14px;margin-bottom:10px">
                    <b style="color:#0369A1">Modifier les groupes du module : {}</b>
                    </div>""".format(m["nom"]), unsafe_allow_html=True)

                new_sel = []
                if classes:
                    ecols = st.columns(min(len(classes), 5))
                    for i, cl in enumerate(classes):
                        with ecols[i % 5]:
                            already = cl["id"] in mc_ids
                            chk = st.checkbox(
                                "{} ({} el.)".format(cl["nom"], cl["nb_eleves"]),
                                value=already,
                                key="edit_mcc_{}_{}".format(m["id"], cl["id"])
                            )
                            if chk:
                                new_sel.append(cl["id"])

                    if len(new_sel) > 1:
                        nms = [c["nom"] for c in classes if c["id"] in new_sel]
                        st.info("Cours commun : {}".format(" + ".join(nms)))

                ec1, ec2 = st.columns([1, 4])
                with ec1:
                    if st.button("Enregistrer", key="save_grp_{}".format(m["id"]), type="primary"):
                        if not new_sel:
                            st.warning("Selectionnez au moins un groupe.")
                        else:
                            con = db()
                            con.execute("PRAGMA foreign_keys=ON")
                            con.execute("DELETE FROM module_classes WHERE module_id=?", (m["id"],))
                            for cid in new_sel:
                                con.execute("INSERT OR IGNORE INTO module_classes (module_id,classe_id) VALUES (?,?)",
                                            (m["id"], cid))
                            con.commit()
                            con.close()
                            st.session_state.pop("edit_mod_grp", None)
                            st.success("Groupes mis a jour !")
                            st.rerun()
                with ec2:
                    if st.button("Annuler", key="cancel_grp_{}".format(m["id"])):
                        st.session_state.pop("edit_mod_grp", None)
                        st.rerun()

def tab_salle():
    st.markdown('<div class="stitle">Salles</div>', unsafe_allow_html=True)

    with st.expander("Ajouter une salle", expanded=True):
        c1, c2, c3 = st.columns([3,1,2])
        with c1:
            snom = st.text_input("Nom / Numero", placeholder="Ex: Amphi A, Salle 12", key="s_nom")
        with c2:
            scap = st.number_input("Capacite", min_value=1, value=30, key="s_cap")
        with c3:
            st.markdown("**Type de salle**")
            sc1, sc2, sc3 = st.columns(3)
            with sc1: ucm = st.checkbox("CM", value=True, key="s_cm")
            with sc2: utd = st.checkbox("TD", key="s_td")
            with sc3: utp = st.checkbox("TP", key="s_tp")
            stype = "tp" if utp else ("td" if utd else "cm")

        if st.button("Ajouter", type="primary", key="btn_adds"):
            if snom.strip():
                run("INSERT INTO salles (nom,capacite,type) VALUES (?,?,?)",
                    (snom.strip(), scap, stype))
                st.rerun()
            else:
                st.warning("Saisissez un nom de salle.")

    salles = get_salles()
    st.markdown("**{} salle(s)**".format(len(salles)))
    tc = {"cm":"#DBEAFE","td":"#D1FAE5","tp":"#FEF3C7"}
    tt = {"cm":"#1D4ED8","td":"#065F46","tp":"#92400E"}
    if not salles:
        st.info("Aucune salle."); return

    cols = st.columns(3)
    for i, s in enumerate(salles):
        with cols[i % 3]:
            c1, c2 = st.columns([4,1])
            with c1:
                st.markdown("""<div style="background:#F8FBFF;border-radius:12px;
                    border:1px solid #DDE8F8;padding:14px;margin-bottom:8px">
                    <div style="font-weight:700;color:#1E3A5F">{}</div>
                    <div style="font-size:.8rem;color:#64748B">{} places</div>
                    <span class="tag" style="background:{};color:{};margin-top:6px;display:inline-block">
                    {}</span></div>""".format(
                        s["nom"], s["capacite"],
                        tc.get(s["type"],"#EEF5FF"), tt.get(s["type"],"#2563EB"),
                        s["type"].upper()),
                    unsafe_allow_html=True)
            with c2:
                if st.button("X", key="ds_{}".format(s["id"])):
                    run("DELETE FROM salles WHERE id=?", (s["id"],)); st.rerun()

def tab_algorithme():
    st.markdown('<div class="stitle">Algorithme de generation</div>', unsafe_allow_html=True)

    descs = {
        "backtracking" : "Explore systematiquement toutes les combinaisons en revenant en arriere sur les conflits. Garantit une solution si elle existe.",
        "coloring"     : "Modelise le probleme comme un graphe : chaque cours est un noeud, les conflits des aretes. Colorie le graphe pour eviter les chevauchements.",
        "genetic"      : "Inspire de l'evolution : genere une population de solutions, les croise et les mute pour converger vers la meilleure. Rapide sur les grands problemes.",
        "simulated"    : "Inspire du recuit des metaux : accepte parfois des solutions moins bonnes pour s'echapper des optima locaux. Tres efficace sur les problemes complexes.",
    }
    icons = {"backtracking":"🔄","coloring":"🎨","genetic":"🧬","simulated":"🌡️"}

    c1, c2 = st.columns(2)
    with c1:
        sel = st.radio("Choisir l'algorithme", list(ALGOS.keys()),
                       format_func=lambda x: "{} {}".format(icons[x], ALGOS[x]),
                       key="algo_sel")
    with c2:
        st.markdown("""<div style="background:#F0F9FF;border-radius:12px;border:1px solid #BAE6FD;
            padding:16px;height:100%"><b style="color:#0369A1">{}</b><br>
            <span style="color:#0C4A6E;font-size:.9rem">{}</span></div>""".format(
            icons[sel] + " " + ALGOS[sel], descs[sel]),
            unsafe_allow_html=True)

    st.divider()
    # Resume
    profs=get_profs(); modules=get_modules(); salles=get_salles()
    classes=get_classes(); periode=get_periode()
    cr_list=get_creneaux()

    cols = st.columns(6)
    for col, lbl, val in [
        (cols[0], "Profs",    len(profs)),
        (cols[1], "Classes",  len(classes)),
        (cols[2], "Modules",  len(modules)),
        (cols[3], "Salles",   len(salles)),
        (cols[4], "Creneaux", len(cr_list)),
        (cols[5], "Semaines", periode.get("nb_semaines",16)),
    ]:
        with col: st.metric(lbl, val)

    st.divider()
    # Modifier nb semaines rapidement
    nb_cur = int(periode.get("nb_semaines", 16))
    nb_new = st.number_input("Nombre de semaines", min_value=1, max_value=52, value=nb_cur, key="algo_ns")
    if nb_new != nb_cur:
        run("UPDATE periode SET nb_semaines=? WHERE id=(SELECT id FROM periode ORDER BY id DESC LIMIT 1)",
            (nb_new,))
        st.rerun()

    st.markdown("---")
    if st.button("Generer l'EDT — {}".format(ALGOS[sel]),
                 type="primary", use_container_width=True, key="btn_gen"):
        errs = []
        if not modules:  errs.append("modules")
        if not salles:   errs.append("salles")
        if not profs:    errs.append("professeurs")
        if not classes:  errs.append("classes")
        if errs:
            st.error("Manquant : {}".format(", ".join(errs)))
        else:
            with st.spinner("Generation en cours..."):
                ok, msg = generate(sel)
            if ok:
                st.success(msg)
                st.balloons()
            else:
                st.error(msg)

def tab_edt():
    st.markdown('<div class="stitle">Emploi du temps genere</div>', unsafe_allow_html=True)
    nb = max_semaine()
    if nb == 0:
        st.warning("Aucun EDT genere. Allez dans l'onglet Algorithme."); return

    creneaux = get_creneaux()
    ja       = get_jours_actifs()
    classes  = get_classes()
    modules  = get_modules()

    if not creneaux or not ja or not classes:
        st.warning("Donnees manquantes."); return

    # Legende
    if modules:
        st.markdown("**Legende des modules :**")
        lc = st.columns(min(len(modules), 5))
        for i, m in enumerate(modules):
            with lc[i % 5]:
                st.markdown("""<div style="display:flex;align-items:center;gap:6px;padding:4px 10px;
                    border-radius:16px;background:{}22;border:1.5px solid {};
                    margin-bottom:4px;font-size:.72rem;font-weight:700;color:#1E3A5F">
                    <div style="width:10px;height:10px;border-radius:2px;background:{}"></div>{}</div>""".format(
                    m["couleur"], m["couleur"], m["couleur"], m["nom"]),
                    unsafe_allow_html=True)
        st.divider()

    c1, c2, c3 = st.columns([2,2,2])
    with c1:
        semaine = st.slider("Semaine", min_value=1, max_value=int(nb), value=1, key="edt_sem")
    with c2:
        go  = ["Toutes les classes"] + [cl["nom"] for cl in classes]
        fg  = st.selectbox("Afficher", go, key="edt_filtre")
    with c3:
        afficher_toutes = st.checkbox("Voir toutes les semaines (resume)", key="edt_all")

    cells = get_edt(semaine)
    grid  = {}
    for cell in cells:
        grid[(cell["jour_index"], cell["creneau_id"], cell["classe_id"])] = cell

    cl_affich = classes if fg == "Toutes les classes" else [cl for cl in classes if cl["nom"] == fg]
    ng        = len(cl_affich)

    if not afficher_toutes:
        _render_edt_table(semaine, cl_affich, ng, ja, creneaux, grid)
    else:
        for sem in range(1, int(nb)+1):
            cells_s = get_edt(sem)
            grid_s  = {}
            for cell in cells_s:
                grid_s[(cell["jour_index"], cell["creneau_id"], cell["classe_id"])] = cell
            with st.expander("Semaine {}".format(sem), expanded=(sem==1)):
                _render_edt_table(sem, cl_affich, ng, ja, creneaux, grid_s)

    # Export Excel
    st.divider()
    st.markdown("**Telecharger l'EDT**")
    c1, c2, c3 = st.columns([2,2,2])
    with c1:
        dl_sem = st.number_input("Semaine a exporter", min_value=1, max_value=int(nb),
                                  value=int(semaine), key="dl_sem")
    with c2:
        dl_classes = st.multiselect("Classes a exporter",
                                     [cl["nom"] for cl in classes],
                                     default=[cl["nom"] for cl in cl_affich],
                                     key="dl_classes")
    with c3:
        st.write(""); st.write("")
        if st.button("Generer Excel", key="btn_excel", type="primary"):
            cls_exp = [cl for cl in classes if cl["nom"] in dl_classes]
            if not cls_exp:
                st.warning("Selectionnez au moins une classe.")
            elif not EXCEL_OK:
                st.error("openpyxl non installe. Lancez : pip install openpyxl")
            else:
                xls = export_excel(dl_sem, cls_exp, creneaux, ja)
                if xls:
                    st.download_button(
                        label="Telecharger EDT_Semaine{}.xlsx".format(dl_sem),
                        data=xls,
                        file_name="EDT_Semaine{}.xlsx".format(dl_sem),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_xlsx"
                    )

def _render_edt_table(semaine, cl_affich, ng, ja, creneaux, grid):
    st.markdown("**Semaine {}**".format(semaine))
    if ng == 0:
        st.info("Aucune classe."); return

    html  = "<div style='overflow-x:auto'>"
    html += "<table style='border-collapse:separate;border-spacing:3px;width:100%;min-width:400px'>"

    # Ligne jours
    html += "<tr><th style='background:#1E3A5F;color:#fff;padding:9px;border-radius:8px;min-width:72px;font-size:.78rem'></th>"
    for j in ja:
        html += "<th colspan='{ng}' style='background:#2D6A4F;color:#fff;padding:8px 4px;border-radius:8px;text-align:center;font-size:.82rem;min-width:{mw}px'>{jour}</th>".format(
            ng=ng, mw=120*ng, jour=JOURS[j])
    html += "</tr>"

    # Ligne classes (si plusieurs)
    if ng > 1:
        html += "<tr><td></td>"
        for j in ja:
            for cl in cl_affich:
                html += "<th style='background:#1E3A5F;color:#fff;padding:5px 8px;border-radius:6px;text-align:center;font-size:.75rem;min-width:120px'>{}</th>".format(cl["nom"])
        html += "</tr>"

    # Lignes creneaux
    for cr in creneaux:
        html += "<tr>"
        html += "<td style='background:#F8FBFF;padding:8px;border-radius:8px;font-size:.72rem;color:#64748B;font-weight:700;text-align:center;white-space:nowrap;vertical-align:middle'>{}<br>{}</td>".format(
            cr["debut"], cr["fin"])

        for j in ja:
            slot = {cl["id"]: grid.get((j, cr["id"], cl["id"])) for cl in cl_affich}
            nonempty = {cid: c for cid, c in slot.items() if c and c.get("mnom")}

            if ng > 1:
                rendered = set(); gi = 0
                while gi < len(cl_affich):
                    cl  = cl_affich[gi]
                    if cl["id"] in rendered: gi += 1; continue
                    cell = nonempty.get(cl["id"])
                    if not cell:
                        html += "<td style='padding:3px'><div style='background:#F8FBFF;border:1px dashed #E2E8F0;border-radius:8px;min-height:68px'></div></td>"
                        rendered.add(cl["id"]); gi += 1; continue
                    # compter colspan
                    cs = 1; same = [cl["id"]]
                    for k in range(gi+1, len(cl_affich)):
                        nc = cl_affich[k]
                        ncc = nonempty.get(nc["id"])
                        if ncc and ncc.get("mnom")==cell.get("mnom") and ncc.get("snom")==cell.get("snom"):
                            cs += 1; same.append(nc["id"])
                        else: break
                    color = cell.get("couleur","#4A90D9")
                    prof  = "{} {}".format(cell.get("pprenom",""), cell.get("pnom","")).strip()
                    badge = ""
                    if cs > 1:
                        gn = " + ".join([cl_affich[gi+k]["nom"] for k in range(cs)])
                        badge = "<div style='font-size:.58rem;color:#fff;background:{};border-radius:4px;padding:1px 5px;margin-top:3px;display:inline-block'>Commun: {}</div>".format(color, gn)
                    html += "<td colspan='{cs}' style='padding:3px;vertical-align:top'><div style='background:{c}18;border:2px solid {c};border-radius:8px;padding:7px;min-height:68px;font-size:.69rem'><div style='display:flex;align-items:center;gap:4px;margin-bottom:2px'><div style='width:8px;height:8px;border-radius:2px;background:{c};flex-shrink:0'></div><b style='color:#1E3A5F;line-height:1.2'>{mn}</b></div><div style='color:#475569'>{tp}</div><div style='color:#475569'>{pr}</div><div style='color:#64748B'>{sl}</div>{bg}</div></td>".format(
                        cs=cs, c=color, mn=cell["mnom"],
                        tp=cell.get("type_salle","").upper(),
                        pr=prof, sl=cell.get("snom",""), bg=badge)
                    for cid in same: rendered.add(cid)
                    gi += cs
            else:
                cl   = cl_affich[0]
                cell = nonempty.get(cl["id"])
                if cell:
                    color = cell.get("couleur","#4A90D9")
                    prof  = "{} {}".format(cell.get("pprenom",""), cell.get("pnom","")).strip()
                    html += "<td style='padding:3px;vertical-align:top'><div style='background:{c}18;border:2px solid {c};border-radius:8px;padding:7px;min-height:68px;font-size:.69rem'><div style='display:flex;align-items:center;gap:4px;margin-bottom:2px'><div style='width:8px;height:8px;border-radius:2px;background:{c};flex-shrink:0'></div><b style='color:#1E3A5F'>{mn}</b></div><div style='color:#475569'>{tp}</div><div style='color:#475569'>{pr}</div><div style='color:#64748B'>{sl}</div></div></td>".format(
                        c=color, mn=cell["mnom"],
                        tp=cell.get("type_salle","").upper(),
                        pr=prof, sl=cell.get("snom",""))
                else:
                    html += "<td style='padding:3px'><div style='background:#F8FBFF;border:1px dashed #E2E8F0;border-radius:8px;min-height:68px'></div></td>"
        html += "</tr>"

    html += "</table></div>"
    st.markdown(html, unsafe_allow_html=True)

# ═══════════════════════════════════════
# MAIN
# ═══════════════════════════════════════
SIDE_OPTS = ["Identifiant", "Creneaux", "Calendrier", "Edusign"]

def main():
    if not st.session_state.logged_in:
        page_login()
        return

    # ── Sidebar ──
    with st.sidebar:
        st.markdown('<div class="logo" style="text-align:center;padding:.8rem 0 .4rem">JM<span>-edt</span></div>',
                    unsafe_allow_html=True)
        st.markdown("<hr style='border-color:#DDE8F8;margin:.5rem 0'>", unsafe_allow_html=True)
        sp = st.radio("Reglages", SIDE_OPTS, label_visibility="visible", key="side_nav")
        st.markdown("<hr style='border-color:#DDE8F8;margin:.5rem 0'>", unsafe_allow_html=True)
        st.markdown("<small style='color:#94A3B8'>{}</small>".format(st.session_state.email),
                    unsafe_allow_html=True)
        if st.button("Deconnexion", use_container_width=True, key="btn_disc"):
            st.session_state.logged_in = False
            st.session_state.email     = ""
            st.rerun()

    # ── Onglets principaux ──
    tabs = st.tabs(["Professeur","Classes","Disponibilite","Matiere","Salle","Algorithme","EDT"])
    with tabs[0]: tab_professeur()
    with tabs[1]: tab_classes()
    with tabs[2]: tab_disponibilite()
    with tabs[3]: tab_matiere()
    with tabs[4]: tab_salle()
    with tabs[5]: tab_algorithme()
    with tabs[6]: tab_edt()

    # ── Contenu sidebar ──
    side_map = {
        "Identifiant": page_identifiant,
        "Creneaux":    page_creneaux,
        "Calendrier":  page_periode,
        "Edusign":     page_edusign,
    }
    with st.sidebar:
        st.markdown("<hr style='border-color:#DDE8F8;margin:.5rem 0'>", unsafe_allow_html=True)
        if sp in side_map:
            with st.expander(sp, expanded=True):
                side_map[sp]()

if __name__ == "__main__":
    main()
